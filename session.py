
import csv  
import binascii
import os
import random
from uuid import uuid4
import pandas as pd
import xlsxwriter
from openpyxl import load_workbook
from datetime import datetime
from secrets import token_hex
from datetime import timedelta
import boto3


#def updatedata():
workbook = load_workbook(filename="/usr/share/mockauth/MockSession.xlsx")
sheet = workbook.active
rowcnt=sheet.max_row

for x in range(1,rowcnt):
     session_token = uuid4()
     nhsid_session_id = uuid4()
     nonce = uuid4()
     session_id = uuid4()
     dt = datetime.now()
     current_time = dt + timedelta(minutes=15)
     expires = int(current_time.timestamp())
     future_time = dt + timedelta(minutes=300)
     reauth=int(future_time.timestamp())
     sheet["A"+str(x+1)] = str(session_token)
     sheet["D"+str(x+1)]=token_hex(32)#str(generate_session_key())
     sheet["E"+str(x+1)]=expires
     sheet["I"+str(x+1)] = str(nhsid_session_id)
     sheet["K"+str(x+1)] = str(nonce)
     sheet["O"+str(x + 1)] = reauth
     sheet["S"+str(x+1)] = str(session_id).strip()

workbook.save(filename="/usr/share/mockauth/MockSession.xlsx")
#Convert Excel to CSV
data=pd.read_excel("/usr/share/mockauth/MockSession.xlsx",sheet_name='Sheet1',dtype=str,index_col=None)
data.to_csv("/usr/share/mockauth/sessiondata.csv",encoding='utf-8',index=False,header=False,line_terminator=None)
data1=pd.read_excel("/usr/share/mockauth/MockSession.xlsx",sheet_name='Sheet1',usecols=['session_token','csrf_token'])
data1.to_csv("/usr/share/mockauth/sessions.csv",encoding='utf-8',index=False,header=True,line_terminator=None)
print("Number of Session Updated in Sheet-"+str(rowcnt))

session=boto3.Session(profile_name='cim-preprod-developer')
s3 = session.client('s3')
s3.upload_file(
     Filename="/home/ec2-user/mockauth/sessiondata.csv",
     Bucket="mocksessionbuc",
     Key="sessiondata.csv",
)
print("User Session created successfully...")

